from __future__ import annotations

import io
import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from db import connect, init_db, now_iso
from exports import build_excel, build_html_report, build_scheduling_csv


class ExportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.path = Path(self.temp.name) / "export.db"
        init_db(self.path)
        with connect(self.path) as db:
            self.eid = self.seed(db)

    def tearDown(self):
        self.temp.cleanup()

    def test_excel_opens_with_engagement_summary_first_and_mirrors_key_positions(self):
        with connect(self.path) as db:
            filename, content = build_excel(db, self.eid)
        wb = load_workbook(io.BytesIO(content), data_only=False)
        self.assertTrue(filename.startswith("Export_Client_P100_"))
        self.assertEqual(wb.sheetnames[:3], ["Engagement Summary", "Weekly Detail", "Adjustment Log"])
        ws = wb["Engagement Summary"]
        self.assertEqual(ws["F2"].value, "Engagement Summary")
        self.assertEqual(ws["B3"].value, "Engagement Name")
        self.assertEqual(ws["C3"].value, "Export Client")
        self.assertEqual(ws["B10"].value, "Engagement Team Summary")
        self.assertEqual(ws["I10"].value, "Effort Summary")
        self.assertEqual(ws["I11"].value, "Phase or Segment")
        self.assertIn("F2:I2", [str(item) for item in ws.merged_cells.ranges])
        wb.close()

    def test_html_report_is_print_ready_and_formats_summary(self):
        with connect(self.path) as db:
            report = build_html_report(db, self.eid, "Delivery is on track.")
        self.assertIn("Engagement Summary", report)
        self.assertIn("Export Client", report)
        self.assertIn("$10,000.00", report)
        self.assertIn("Engagement Team Summary", report)
        self.assertIn("Effort Summary", report)
        self.assertIn("@page", report)
        self.assertIn("Delivery is on track.", report)

    def test_scheduling_csv_uses_current_monday_cutoff_and_null_forecast_fallback(self):
        with connect(self.path) as db:
            member = db.execute("SELECT id FROM team_members WHERE engagement_id=?", (self.eid,)).fetchone()["id"]
            phase = db.execute("SELECT id FROM phases WHERE engagement_id=?", (self.eid,)).fetchone()["id"]
            db.execute("""INSERT INTO phase_person_weeks
                (phase_id,team_member_id,week_start_date,budgeted_hours,forecasted_hours)
                VALUES (?,?,?,?,?)""", (phase, member, "2026-08-03", 5, 7))
            db.execute("""INSERT INTO phase_person_weeks
                (phase_id,team_member_id,week_start_date,budgeted_hours,forecasted_hours)
                VALUES (?,?,?,?,?)""", (phase, member, "2026-08-10", 8, None))
            db.execute("""INSERT INTO phase_person_weeks
                (phase_id,team_member_id,week_start_date,budgeted_hours,forecasted_hours)
                VALUES (?,?,?,?,?)""", (phase, member, "2026-08-17", 10, 0))
            filename, content = build_scheduling_csv(db, self.eid, today=date(2026, 8, 13))
        rows = list(csv.reader(io.StringIO(content.decode("utf-8"))))
        self.assertIn("_scheduling_", filename)
        self.assertEqual(rows[0], ["Worker", "Role", "2026-08-10", "2026-08-17"])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], "Smith, Jane")
        self.assertEqual(rows[1][2], "8.00")
        self.assertEqual(rows[1][3], "0.00")

    def seed(self, db):
        eid=db.execute("""INSERT INTO engagements
            (engagement_code,client_name,complexity_mode,status,engagement_lead,created_at,updated_at)
            VALUES ('P100','Export Client','complex','active','Lead, Alex',?,?)""",
            (now_iso(),now_iso())).lastrowid
        pid=db.execute("""INSERT INTO phases
            (engagement_id,phase_name,phase_code,sow_fees,created_at) VALUES (?,'Phase 1','A',10000,?)""",
            (eid,now_iso())).lastrowid
        member=db.execute("""INSERT INTO team_members
            (engagement_id,name,role,internal_rate,engagement_rate,contract_rate)
            VALUES (?,'Smith, Jane','Manager FY26',350,300,280)""",(eid,)).lastrowid
        db.execute("""INSERT INTO phase_person_weeks
            (phase_id,team_member_id,week_start_date,budgeted_hours) VALUES (?,?,?,40)""",
            (pid,member,"2026-07-06"))
        snapshot=db.execute("""INSERT INTO weekly_snapshots
            (engagement_id,week_end_date,imported_at,row_count) VALUES (?,'2026-07-12',?,1)""",
            (eid,now_iso())).lastrowid
        db.execute("""INSERT INTO time_entries
            (snapshot_id,engagement_id,transaction_id,worker_name,week_end_date,hours,
             fees_std_rate,fees_contract_rate,matched_phase_id)
            VALUES (?,?,'T1','Smith, Jane','2026-07-12',10,3500,2800,?)""",
            (snapshot,eid,pid))
        return eid


if __name__ == "__main__":
    unittest.main()
