from __future__ import annotations

import io
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from calculations import week_monday
from db import connect, init_db, now_iso
from importers import (
    covered_period_from_text,
    EXPECTED_COLUMNS,
    excel_serial_to_iso,
    parse_text_export,
    parse_xlsx_export,
    preview_rows,
    suggest_from_memo,
)

ROW_DEFAULT_WEEK_START = week_monday(excel_serial_to_iso("45129"))


class ImporterTests(unittest.TestCase):
    def test_covered_period_is_read_from_cognos_preamble(self):
        self.assertEqual(covered_period_from_text("Time detail From: 07/01/2026 to 07/31/2026"),
                         ("2026-07-01", "2026-07-31"))
    def test_excel_serial_date_conversion(self):
        self.assertEqual(excel_serial_to_iso("45123"), "2023-07-16")

    def test_parse_xlsx_export_finds_header_below_preamble(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Time Detail"
        sheet["B2"] = "Time and Cost Detail for Customer: Sample"
        sheet["B4"] = "Project: P100 - Sample"
        sheet["B6"] = "From: 4/1/2026 to 6/17/2026"
        sheet["B8"] = "Confidential data for internal use only."
        for col, header in enumerate(EXPECTED_COLUMNS, 2):
            sheet.cell(row=11, column=col, value=header)
        values = [
            "T1",
            "W1",
            "Smith, Jane",
            "Manager",
            "CONSULT RISK",
            "MRLR",
            "45123",
            "45129",
            "2026-06",
            "P100",
            "Project",
            "P100",
            "Phase",
            "Task",
            "Remote",
            "Billable",
            1.5,
            300,
            250,
            "Memo",
        ]
        for col, value in enumerate(values, 2):
            sheet.cell(row=12, column=col, value=value)
        sheet["B13"] = "Overall- Summary"
        sheet["R13"] = 1.5
        output = io.BytesIO()
        workbook.save(output)

        rows = parse_xlsx_export(output.getvalue())

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Transaction ID"], "T1")
        self.assertEqual(rows[0]["Date"], "2023-07-16")
        self.assertEqual(rows[0]["Worker"], "Smith, Jane")

    def test_preview_flags_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """
                    INSERT INTO engagements (
                      engagement_code, client_name, status, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,is_default) VALUES (?,'General',1)",
                    (engagement_id,),
                )
                conn.execute(
                    """
                    INSERT INTO team_members (engagement_id, name, role)
                    VALUES (?, ?, ?)
                    """,
                    (engagement_id, "Smith, Jane", "Manager"),
                )
                conn.execute(
                    """
                    INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count)
                    VALUES (?, ?, ?, ?)
                    """,
                    (engagement_id, "2026-06-19", now_iso(), 1),
                )
                snapshot_id = conn.execute("SELECT id FROM weekly_snapshots").fetchone()["id"]
                conn.execute(
                    """
                    INSERT INTO time_entries (snapshot_id, engagement_id, transaction_id)
                    VALUES (?, ?, ?)
                    """,
                    (snapshot_id, engagement_id, "DUP"),
                )
                rows = [
                    self._row("OK", "Smith, Jane", "P100", 1),
                    self._row("MIS", "Smith, Jane", "P200", 1),
                    self._row("UNK", "Unknown, Worker", "P100", 1),
                    self._row("ZERO", "Smith, Jane", "P100", 0),
                    self._row("DUP", "Smith, Jane", "P100", 1),
                ]
                preview = preview_rows(conn, engagement_id, rows)

        flags = {row["transaction_id"]: row["flag"] for row in preview["rows"]}
        self.assertIsNone(flags["OK"])
        self.assertEqual(flags["MIS"], "project_mismatch")
        self.assertEqual(flags["UNK"], "worker_unknown")
        self.assertEqual(flags["ZERO"], "zero_hours")
        self.assertIsNone(flags["DUP"])
        self.assertEqual(preview["summary"]["to_import"], 5)
        duplicate = next(row for row in preview["rows"] if row["transaction_id"] == "DUP")
        self.assertEqual(duplicate["reconciliation_action"], "update")
        mismatch = next(row for row in preview["rows"] if row["transaction_id"] == "MIS")
        self.assertTrue(mismatch["included"])

    def test_task_desc_matches_phase_when_phase_desc_unmatched(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """INSERT INTO engagements (engagement_code, client_name, status, complexity_mode,
                       created_at, updated_at) VALUES (?, ?, ?, 'complex', ?, ?)""",
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                phase1_id = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'ETL Build','ETL')",
                    (engagement_id,),
                ).lastrowid
                phase2_id = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'Migration','MIG')",
                    (engagement_id,),
                ).lastrowid
                conn.execute(
                    "INSERT INTO team_members (engagement_id, name, role) VALUES (?, ?, ?)",
                    (engagement_id, "Smith, Jane", "Manager"),
                )
                conn.execute(
                    "INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count) "
                    "VALUES (?, ?, ?, ?)", (engagement_id, "2026-06-19", now_iso(), 3),
                )
                rows = [
                    self._row("DIRECT", "Smith, Jane", "P100", 1, phase_desc="ETL"),
                    self._row("TASKMATCH", "Smith, Jane", "P100", 1, phase_desc="Not A Phase", task_desc="MIG"),
                    self._row("NOMATCH", "Smith, Jane", "P100", 1, phase_desc="Not A Phase", task_desc="Also Not A Phase"),
                ]
                preview = preview_rows(conn, engagement_id, rows)

        by_id = {row["transaction_id"]: row for row in preview["rows"]}
        self.assertEqual(by_id["DIRECT"]["matched_phase_id"], phase1_id)
        self.assertEqual(by_id["DIRECT"]["allocation_method"], "direct_match")
        self.assertEqual(by_id["TASKMATCH"]["matched_phase_id"], phase2_id)
        self.assertEqual(by_id["TASKMATCH"]["allocation_method"], "task_match")
        self.assertIsNone(by_id["NOMATCH"]["matched_phase_id"])
        self.assertIsNone(by_id["NOMATCH"]["allocation_method"])
        self.assertIn("unmatched_phase", by_id["NOMATCH"]["flags"])

    def test_sticky_rule_auto_applies_and_suppresses_unmatched_flag(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """INSERT INTO engagements (engagement_code, client_name, status, complexity_mode,
                       created_at, updated_at) VALUES (?, ?, ?, 'complex', ?, ?)""",
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                phase_id = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'ETL Build','ETL')",
                    (engagement_id,),
                ).lastrowid
                member_id = conn.execute(
                    "INSERT INTO team_members (engagement_id, name, role) VALUES (?, ?, ?)",
                    (engagement_id, "Smith, Jane", "Manager"),
                ).lastrowid
                conn.execute(
                    "INSERT INTO allocation_rules (engagement_id,team_member_id,phase_id,created_at) "
                    "VALUES (?,?,?,?)", (engagement_id, member_id, phase_id, now_iso()),
                )
                conn.execute(
                    "INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count) "
                    "VALUES (?, ?, ?, ?)", (engagement_id, "2026-06-19", now_iso(), 1),
                )
                rows = [self._row("RULED", "Smith, Jane", "P100", 1, phase_desc="Not A Phase", task_desc="Also Not A Phase")]
                preview = preview_rows(conn, engagement_id, rows)

        row = preview["rows"][0]
        self.assertEqual(row["matched_phase_id"], phase_id)
        self.assertEqual(row["allocation_method"], "sticky_rule")
        self.assertNotIn("unmatched_phase", row["flags"])
        self.assertIsNone(row["flag"])

    def test_single_phase_budget_auto_resolves_without_flagging(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """INSERT INTO engagements (engagement_code, client_name, status, complexity_mode,
                       created_at, updated_at) VALUES (?, ?, ?, 'complex', ?, ?)""",
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                governance_id = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'AI Governance','GOV')",
                    (engagement_id,),
                ).lastrowid
                conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'Migration','MIG')",
                    (engagement_id,),
                )
                member_id = conn.execute(
                    "INSERT INTO team_members (engagement_id, name, role) VALUES (?, ?, ?)",
                    (engagement_id, "Minard, Corey", "Partner"),
                ).lastrowid
                week_start = ROW_DEFAULT_WEEK_START
                conn.execute(
                    "INSERT INTO phase_person_weeks (phase_id,team_member_id,week_start_date,budgeted_hours) "
                    "VALUES (?,?,?,10)", (governance_id, member_id, week_start),
                )
                conn.execute(
                    "INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count) "
                    "VALUES (?, ?, ?, ?)", (engagement_id, "2026-06-19", now_iso(), 1),
                )
                rows = [self._row("SOLO", "Minard, Corey", "P100", 1, phase_desc="Not A Phase", task_desc="Also Not A Phase")]
                preview = preview_rows(conn, engagement_id, rows)

        row = preview["rows"][0]
        self.assertEqual(row["matched_phase_id"], governance_id)
        self.assertEqual(row["allocation_method"], "single_phase_budget")
        self.assertNotIn("unmatched_phase", row["flags"])
        self.assertIsNone(row["flag"])

    def test_single_phase_budget_auto_resolves_even_outside_staffing_window(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """INSERT INTO engagements (engagement_code, client_name, status, complexity_mode,
                       created_at, updated_at) VALUES (?, ?, ?, 'complex', ?, ?)""",
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                governance_id = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'AI Governance','GOV')",
                    (engagement_id,),
                ).lastrowid
                conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'Migration','MIG')",
                    (engagement_id,),
                )
                member_id = conn.execute(
                    "INSERT INTO team_members (engagement_id, name, role) VALUES (?, ?, ?)",
                    (engagement_id, "Minard, Corey", "Partner"),
                ).lastrowid
                # Budgeted week is far outside the +/-7-day staffing_match window around the
                # time entry's week below, but Governance is this member's only budgeted phase
                # anywhere in the engagement — that alone should resolve the match.
                conn.execute(
                    "INSERT INTO phase_person_weeks (phase_id,team_member_id,week_start_date,budgeted_hours) "
                    "VALUES (?,?,?,10)", (governance_id, member_id, "2026-01-05"),
                )
                conn.execute(
                    "INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count) "
                    "VALUES (?, ?, ?, ?)", (engagement_id, "2026-06-19", now_iso(), 1),
                )
                row = self._row("SOLO2", "Minard, Corey", "P100", 1, phase_desc="Not A Phase", task_desc="Also Not A Phase")
                preview = preview_rows(conn, engagement_id, [row])

        row = preview["rows"][0]
        self.assertEqual(row["matched_phase_id"], governance_id)
        self.assertEqual(row["allocation_method"], "single_phase_budget")
        self.assertNotIn("unmatched_phase", row["flags"])
        self.assertIsNone(row["flag"])

    def test_multiple_staffing_candidates_still_flag_unmatched(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_file = Path(temp_dir) / "test.db"
            init_db(db_file)
            with connect(db_file) as conn:
                engagement_id = conn.execute(
                    """INSERT INTO engagements (engagement_code, client_name, status, complexity_mode,
                       created_at, updated_at) VALUES (?, ?, ?, 'complex', ?, ?)""",
                    ("P100", "Client", "active", now_iso(), now_iso()),
                ).lastrowid
                phase_a = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'AI Governance','GOV')",
                    (engagement_id,),
                ).lastrowid
                phase_b = conn.execute(
                    "INSERT INTO phases (engagement_id,phase_name,phase_code) VALUES (?,'Migration','MIG')",
                    (engagement_id,),
                ).lastrowid
                member_id = conn.execute(
                    "INSERT INTO team_members (engagement_id, name, role) VALUES (?, ?, ?)",
                    (engagement_id, "Smith, Jane", "Manager"),
                ).lastrowid
                week_start = ROW_DEFAULT_WEEK_START
                conn.execute(
                    "INSERT INTO phase_person_weeks (phase_id,team_member_id,week_start_date,budgeted_hours) "
                    "VALUES (?,?,?,10)", (phase_a, member_id, week_start),
                )
                conn.execute(
                    "INSERT INTO phase_person_weeks (phase_id,team_member_id,week_start_date,budgeted_hours) "
                    "VALUES (?,?,?,10)", (phase_b, member_id, week_start),
                )
                conn.execute(
                    "INSERT INTO weekly_snapshots (engagement_id, week_end_date, imported_at, row_count) "
                    "VALUES (?, ?, ?, ?)", (engagement_id, "2026-06-19", now_iso(), 1),
                )
                rows = [self._row("BOTH", "Smith, Jane", "P100", 1, phase_desc="Not A Phase", task_desc="Also Not A Phase")]
                preview = preview_rows(conn, engagement_id, rows)

        row = preview["rows"][0]
        self.assertIsNone(row["matched_phase_id"])
        self.assertIsNone(row["allocation_method"])
        self.assertIn("unmatched_phase", row["flags"])

    def test_suggest_from_memo_matches_phase_code_or_name(self):
        phases = [
            {"id": 1, "phase_name": "ETL Code Development", "phase_code": "ETL"},
            {"id": 2, "phase_name": "Migration", "phase_code": "MIG"},
        ]
        by_code = suggest_from_memo("worked on ETL fixes this week", phases)
        self.assertEqual(by_code["phase_id"], 1)
        by_name = suggest_from_memo("helping with migration cutover", phases)
        self.assertEqual(by_name["phase_id"], 2)
        self.assertIsNone(suggest_from_memo("general catch-up", phases))
        self.assertIsNone(suggest_from_memo("", phases))

    def test_parse_tab_delimited_export(self):
        line = "\t".join(EXPECTED_COLUMNS)
        values = ["T1", "W1", "Smith, Jane", "Manager", "", "", "45123", "45129", "2026-06", "P100", "Project", "", "Phase", "Task", "Remote", "Billable", "1.5", "300", "250", "Memo"]
        rows = parse_text_export(line + "\n" + "\t".join(values))
        self.assertEqual(rows[0]["Date"], "2023-07-16")
        self.assertEqual(rows[0]["Worker"], "Smith, Jane")

    def _row(self, transaction_id: str, worker: str, project_id: str, hours: float,
             phase_desc: str = "Phase", task_desc: str = "Task"):
        return {
            "Transaction ID": transaction_id,
            "Worker ID": "W1",
            "Worker": worker,
            "Title": "Manager",
            "Date": "45123",
            "Week End Date": "45129",
            "Financial Period": "2026-06",
            "Project ID": project_id,
            "Project": "Project",
            "Phase Desc": phase_desc,
            "Task Desc": task_desc,
            "Work Loc": "Remote",
            "Billing Status": "Billable",
            "Hours": hours,
            "Fees @ Std Rate": 300,
            "Fees @ Contract Rate": 250,
            "Memo": "",
        }


if __name__ == "__main__":
    unittest.main()
