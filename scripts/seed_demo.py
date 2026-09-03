from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_DIR = ROOT.parent / "reference" / "B2A Examples"
sys.path.insert(0, str(ROOT))

from db import connect, init_db, now_iso  # noqa: E402


def normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def text(value: Any) -> str:
    return str(value or "").strip()


def as_float(value: Any) -> float:
    if value in (None, "", " ", "#VALUE!"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def iso(value: Any) -> str | None:
    parsed = as_date(value)
    return parsed.isoformat() if parsed else None


def fallback_code(name: str, path: Path) -> str:
    slug = re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")
    if not slug:
        slug = re.sub(r"[^A-Z0-9]+", "-", path.stem.upper()).strip("-")
    return slug[:40] or f"EXAMPLE-{path.stem[:20].upper()}"


def load_rows(workbook, sheet_name: str, max_cols: int = 20) -> list[list[Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row[:max_cols]))
    return rows


def find_metadata(rows: list[list[Any]], label: str) -> Any:
    target = normalize(label)
    for row in rows[:12]:
        for index, cell in enumerate(row):
            if normalize(cell) == target:
                for candidate in row[index + 1 : index + 4]:
                    if candidate not in (None, ""):
                        return candidate
    return None


def find_header_index(rows: list[list[Any]], required: set[str]) -> int | None:
    for index, row in enumerate(rows):
        present = {normalize(cell) for cell in row if normalize(cell)}
        if required.issubset(present):
            return index
    return None


def build_header_map(headers: list[Any]) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        mapping[normalize(header)].append(index)
    return dict(mapping)


def parse_phase_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = find_header_index(rows, {"phase or segment", "hours"})
    if header_index is None:
        return []
    headers = build_header_map(rows[header_index])
    phase_col = headers["phase or segment"][0]
    hours_cols = headers.get("hours", [])
    budget_hours_col = hours_cols[0] if hours_cols else None
    actual_hours_col = hours_cols[1] if len(hours_cols) > 1 else None
    sow_col = headers.get("sow fees", [None])[0]
    eng_fee_cols = headers.get("eng. fees", [])
    budget_fee_col = eng_fee_cols[0] if eng_fee_cols else None
    phases: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        name = text(row[phase_col] if phase_col < len(row) else "")
        if not name:
            if phases:
                break
            continue
        if normalize(name) == "total":
            break
        budget_hours = as_float(row[budget_hours_col]) if budget_hours_col is not None and budget_hours_col < len(row) else 0
        actual_hours = as_float(row[actual_hours_col]) if actual_hours_col is not None and actual_hours_col < len(row) else 0
        sow_fees = as_float(row[sow_col]) if sow_col is not None and sow_col < len(row) else 0
        if sow_fees == 0 and budget_fee_col is not None and budget_fee_col < len(row):
            sow_fees = as_float(row[budget_fee_col])
        phases.append(
            {
                "phase_name": name,
                "phase_code": name,
                "budgeted_hours": budget_hours,
                "actual_hours": actual_hours,
                "sow_fees": sow_fees,
            }
        )
    return phases


def parse_team_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = None
    for index, row in enumerate(rows):
        normalized = [normalize(cell) for cell in row]
        if "name" in normalized and "project role" in normalized and any(
            item in normalized for item in {"budgeted hours", "hours (budget)"}
        ):
            header_index = index
            break
    if header_index is None:
        return []
    headers = build_header_map(rows[header_index])
    name_col = headers["name"][0]
    role_col = headers["project role"][0]
    internal_col = headers.get("int/std. rates", headers.get("internal / standard rate", [None]))[0]
    engagement_col = headers.get("eng. rates", headers.get("engagement rate", [None]))[0]
    contract_col = headers.get("contract rate", headers.get("contract rates", headers.get("ext. rates", headers.get("external rate", [None]))))[0]
    budget_col = headers.get("budgeted hours", headers.get("hours (budget)", [None]))[0]
    actual_col = headers.get("forecasted actuals hours", headers.get("actuals/current plan hours", headers.get("hours (actual/current)", [None])))[0]
    members: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        name = text(row[name_col] if name_col < len(row) else "")
        role = text(row[role_col] if role_col < len(row) else "")
        if not name and not role:
            if members:
                break
            continue
        budgeted_hours = as_float(row[budget_col]) if budget_col is not None and budget_col < len(row) else 0
        actual_hours = as_float(row[actual_col]) if actual_col is not None and actual_col < len(row) else 0
        if not name or (not role and budgeted_hours == 0 and actual_hours == 0):
            if members:
                break
            continue
        members.append(
            {
                "name": name,
                "role": role or "Unspecified",
                "internal_rate": as_float(row[internal_col]) if internal_col is not None and internal_col < len(row) else 0,
                "engagement_rate": as_float(row[engagement_col]) if engagement_col is not None and engagement_col < len(row) else 0,
                "contract_rate": as_float(row[contract_col]) if contract_col is not None and contract_col < len(row) else 0,
                "budgeted_hours": budgeted_hours,
                "actual_hours": actual_hours,
            }
        )
    return members


def parse_as_of_week(rows: list[list[Any]]) -> int:
    for row in rows:
        normalized = [normalize(cell) for cell in row]
        for index, cell in enumerate(normalized):
            if cell == "as of week":
                if index + 1 < len(row):
                    return max(0, int(as_float(row[index + 1])))
    return 0


def monday(value: Any) -> date | None:
    parsed = as_date(value)
    if not parsed:
        return None
    return parsed - timedelta(days=parsed.weekday())


def select_planner(workbook) -> tuple[str, list[list[Any]], list[dict[str, Any]]]:
    candidates = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("Planner-Tracker"):
            continue
        rows = load_rows(workbook, sheet_name, max_cols=100)
        phases = parse_phase_rows(rows)
        score = (
            sum(max(as_float(phase["sow_fees"]), 0) for phase in phases),
            sum(max(as_float(phase["actual_hours"]), 0) for phase in phases),
            sum(max(as_float(phase["budgeted_hours"]), 0) for phase in phases),
        )
        candidates.append((score, sheet_name, rows, phases))
    if not candidates:
        raise ValueError("Workbook does not contain a Planner-Tracker sheet")
    _score, sheet_name, rows, phases = max(candidates, key=lambda item: item[0])
    return sheet_name, rows, phases


def remove_rollup_phases(phases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Avoid counting a workbook's total row as both a phase and its children."""
    if len(phases) < 2:
        return phases
    for phase in phases:
        others = [candidate for candidate in phases if candidate is not phase]
        other_budget = sum(as_float(candidate["budgeted_hours"]) for candidate in others)
        other_actual = sum(as_float(candidate["actual_hours"]) for candidate in others)
        if (
            abs(as_float(phase["budgeted_hours"]) - other_budget) < 0.01
            and as_float(phase["actual_hours"]) > 0
            and abs(other_actual) < 0.01
            and as_float(phase["sow_fees"]) > 0
        ):
            return [phase]
    return phases


def parse_phase_plans(
    rows: list[list[Any]], phases: list[dict[str, Any]], team: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    member_names = {normalize(member["name"]): member["name"] for member in team}
    phase_names = {normalize(phase["phase_name"]) for phase in phases}
    blocks: list[tuple[int, int, dict[str, Any], list[tuple[int, date]]]] = []
    for phase in phases:
        marker = None
        header = None
        dates: list[tuple[int, date]] = []
        for row_index, row in enumerate(rows):
            if row_index < 20 or not any(normalize(cell) == normalize(phase["phase_name"]) for cell in row[:5]):
                continue
            for candidate_index in range(row_index, min(row_index + 5, len(rows))):
                candidate_dates = [
                    (column_index, monday(value))
                    for column_index, value in enumerate(rows[candidate_index])
                    if monday(value)
                ]
                if candidate_dates and any(normalize(value) == "hours" for value in rows[candidate_index]):
                    marker = row_index
                    header = candidate_index
                    dates = [(column, value) for column, value in candidate_dates if value]
                    break
            if header is not None:
                break
        if marker is None or header is None:
            continue
        blocks.append((marker, header, phase, dates))

    blocks.sort(key=lambda item: item[0])
    plans: list[dict[str, Any]] = []
    for block_index, (_marker, header, phase, dates) in enumerate(blocks):
        end = blocks[block_index + 1][0] if block_index + 1 < len(blocks) else len(rows)
        for candidate_index in range(header + 1, end):
            if any(normalize(value) == "week-->" for value in rows[candidate_index]):
                end = candidate_index
                break
        row_index = header + 1
        while row_index < end:
            row = rows[row_index]
            normalized = [normalize(value) for value in row]
            if "budget" not in normalized:
                row_index += 1
                continue
            source_name = next(
                (member_names[value] for value in normalized if value in member_names), None
            )
            if not source_name:
                row_index += 1
                continue
            current_row = rows[row_index + 1] if row_index + 1 < end else []
            if "actuals/current plan" not in {normalize(value) for value in current_row}:
                current_row = []
            weeks = []
            for column, week_start in dates:
                budget_value = row[column] if column < len(row) else None
                current_value = current_row[column] if column < len(current_row) else None
                weeks.append({
                    "week_start_date": week_start.isoformat(),
                    "budgeted_hours": None if budget_value in (None, "") else as_float(budget_value),
                    "current_plan_hours": None if current_value in (None, "") else as_float(current_value),
                })
            plans.append({
                "phase_name": phase["phase_name"],
                "member_name": source_name,
                "weeks": weeks,
            })
            row_index += 2
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for plan in plans:
        key = (normalize(plan["phase_name"]), normalize(plan["member_name"]))
        target = merged.setdefault(key, {
            "phase_name": plan["phase_name"],
            "member_name": plan["member_name"],
            "weeks": {},
        })
        for week in plan["weeks"]:
            existing = target["weeks"].setdefault(week["week_start_date"], {
                "week_start_date": week["week_start_date"],
                "budgeted_hours": None,
                "current_plan_hours": None,
            })
            for field in ("budgeted_hours", "current_plan_hours"):
                value = week[field]
                if value is not None:
                    existing[field] = as_float(existing[field]) + as_float(value)
    return [
        {
            "phase_name": plan["phase_name"],
            "member_name": plan["member_name"],
            "weeks": [plan["weeks"][week] for week in sorted(plan["weeks"])],
        }
        for plan in merged.values()
    ]


def latest_source_week(workbook) -> date | None:
    latest = None
    for sheet_name in ("Cognos", "Time Data", "Time Detail"):
        if sheet_name not in workbook.sheetnames:
            continue
        rows = load_rows(workbook, sheet_name, max_cols=45)
        header_index = find_header_index(rows, {"week end date"})
        if header_index is None:
            continue
        headers = build_header_map(rows[header_index])
        column = headers["week end date"][0]
        for row in rows[header_index + 1 :]:
            if column >= len(row):
                continue
            parsed = monday(row[column])
            if parsed and (latest is None or parsed > latest):
                latest = parsed
    return latest


def reporting_week_count(
    workbook, selected_rows: list[list[Any]], actual_rows: list[list[Any]], plans: list[dict[str, Any]]
) -> int:
    weeks = sorted({
        date.fromisoformat(week["week_start_date"])
        for plan in plans for week in plan["weeks"]
    })
    if not weeks:
        return 0
    marked_week = parse_as_of_week(actual_rows)
    if marked_week and not any("#value" in normalize(value) for row in actual_rows for value in row):
        return min(marked_week, len(weeks))
    cutoff = monday(find_metadata(selected_rows, "Last day of the week")) or latest_source_week(workbook)
    if cutoff:
        return sum(1 for week in weeks if week <= cutoff)
    return 0


def parse_example(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        base_planner_rows = load_rows(workbook, "Planner-Tracker", max_cols=100)
        planner_sheet, planner_rows, phases = select_planner(workbook)
        summary_rows = load_rows(workbook, "Engagement Summary", max_cols=20)
        actual_rows = load_rows(workbook, "Actuals to Budget", max_cols=12)
        name = text(find_metadata(planner_rows, "Engagement Name") or find_metadata(base_planner_rows, "Engagement Name")) or path.stem
        code = text(find_metadata(planner_rows, "Engagement Number") or find_metadata(base_planner_rows, "Engagement Number")) or fallback_code(name, path)
        lead = text(find_metadata(planner_rows, "Engagement Lead") or find_metadata(base_planner_rows, "Engagement Lead")) or "Unassigned"
        phases = remove_rollup_phases(phases)
        team = parse_team_rows(planner_rows)
        if not team and summary_rows:
            team = parse_team_rows(summary_rows)
        if not phases:
            phases = [{"phase_name": "General", "phase_code": "General", "budgeted_hours": 0, "actual_hours": 0, "sow_fees": 0}]
        plans = parse_phase_plans(planner_rows, phases, team)
        plan_weeks = sorted({week["week_start_date"] for plan in plans for week in plan["weeks"]})
        first_monday = plan_weeks[0] if plan_weeks else (
            iso(find_metadata(planner_rows, "First Monday") or find_metadata(base_planner_rows, "First Monday"))
            or date(2026, 1, 5).isoformat()
        )
        completed_weeks = reporting_week_count(workbook, planner_rows, actual_rows, plans)
        return {
            "path": path,
            "planner_sheet": planner_sheet,
            "client_name": name,
            "engagement_code": code,
            "engagement_lead": lead,
            "first_monday": first_monday,
            "duration_weeks": max(1, len(plan_weeks)),
            "completed_weeks": completed_weeks,
            "phases": phases,
            "team": team,
            "plans": plans,
        }
    finally:
        workbook.close()


def insert_engagement(conn, data: dict[str, Any]) -> int:
    now = now_iso()
    cursor = conn.execute(
        """
        INSERT INTO engagements (
          engagement_code, client_name, engagement_type, complexity_mode, model_type,
          model_vendor, engagement_lead, first_monday, duration_weeks, status,
          c360_used, c360_amount, bima_amount, created_at, updated_at
        ) VALUES (?, ?, ?, 'complex', ?, ?, ?, ?, ?, 'active', 0, 0, 0, ?, ?)
        """,
        (
            data["engagement_code"],
            data["client_name"],
            "B2A Example",
            "Budget to Actual",
            data["path"].name,
            data["engagement_lead"],
            data["first_monday"],
            data["duration_weeks"],
            now,
            now,
        ),
    )
    return int(cursor.lastrowid)


def insert_phase(conn, engagement_id: int, phase: dict[str, Any], order: int) -> int:
    cursor = conn.execute(
        """
        INSERT INTO phases (
          engagement_id, phase_name, phase_code, sow_fees, sort_order, is_default, created_at
        ) VALUES (?, ?, ?, ?, ?, 0, ?)
        """,
        (
            engagement_id,
            phase["phase_name"],
            phase["phase_code"],
            phase["sow_fees"],
            order,
            now_iso(),
        ),
    )
    return int(cursor.lastrowid)


def insert_member(conn, engagement_id: int, member: dict[str, Any]) -> int:
    cursor = conn.execute(
        """
        INSERT INTO team_members (
          engagement_id, name, role, is_offshore, is_active, internal_rate,
          engagement_rate, contract_rate, dte_rate, created_at
        ) VALUES (?, ?, ?, 0, 1, ?, ?, ?, ?, ?)
        """,
        (
            engagement_id,
            member["name"],
            member["role"],
            member["internal_rate"],
            member["engagement_rate"] or member["internal_rate"],
            member["contract_rate"] or member["engagement_rate"] or member["internal_rate"],
            member["engagement_rate"] or member["internal_rate"],
            now_iso(),
        ),
    )
    return int(cursor.lastrowid)


def insert_source_plans(
    conn, data: dict[str, Any], phase_ids: list[int], member_ids: list[int]
) -> None:
    phase_by_name = {
        normalize(phase["phase_name"]): phase_id
        for phase, phase_id in zip(data["phases"], phase_ids)
    }
    member_by_name = {
        normalize(member["name"]): member_id
        for member, member_id in zip(data["team"], member_ids)
    }
    all_weeks = sorted({
        week["week_start_date"]
        for plan in data["plans"] for week in plan["weeks"]
    })
    completed = set(all_weeks[: data["completed_weeks"]])
    for plan in data["plans"]:
        phase_id = phase_by_name.get(normalize(plan["phase_name"]))
        member_id = member_by_name.get(normalize(plan["member_name"]))
        if not phase_id or not member_id:
            continue
        for week in plan["weeks"]:
            current = week["current_plan_hours"]
            forecast = None if week["week_start_date"] in completed else current
            conn.execute(
                """
                INSERT INTO phase_person_weeks (
                  phase_id, team_member_id, week_start_date, budgeted_hours, forecasted_hours
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (
                    phase_id,
                    member_id,
                    week["week_start_date"],
                    as_float(week["budgeted_hours"]),
                    forecast,
                ),
            )


def insert_source_actuals(
    conn, engagement_id: int, data: dict[str, Any], phase_ids: list[int], member_ids: list[int]
) -> None:
    phase_by_name = {
        normalize(phase["phase_name"]): phase_id
        for phase, phase_id in zip(data["phases"], phase_ids)
    }
    member_by_name = {
        normalize(member["name"]): (member, member_id)
        for member, member_id in zip(data["team"], member_ids)
    }
    all_weeks = sorted({
        week["week_start_date"]
        for plan in data["plans"] for week in plan["weeks"]
    })
    completed = all_weeks[: data["completed_weeks"]]
    for week_index, week_start_value in enumerate(completed, start=1):
        week_start = date.fromisoformat(week_start_value)
        week_end = week_start + timedelta(days=5)
        snapshot_id = conn.execute(
            """
            INSERT INTO weekly_snapshots (
              engagement_id, week_end_date, imported_at, row_count, notes
            ) VALUES (?, ?, ?, 0, ?)
            """,
            (
                engagement_id,
                week_end.isoformat(),
                now_iso(),
                f"Loaded from source workbook: {data['path'].name}",
            ),
        ).lastrowid
        rows_inserted = 0
        for plan_index, plan in enumerate(data["plans"], start=1):
            phase_id = phase_by_name.get(normalize(plan["phase_name"]))
            member_record = member_by_name.get(normalize(plan["member_name"]))
            if not phase_id or not member_record:
                continue
            member, member_id = member_record
            source_week = next(
                (week for week in plan["weeks"] if week["week_start_date"] == week_start_value), None
            )
            hours = as_float(source_week["current_plan_hours"]) if source_week else 0
            if hours <= 0:
                continue
            contract_rate = member["contract_rate"] or member["engagement_rate"] or member["internal_rate"]
            standard_rate = member["internal_rate"] or contract_rate
            transaction_id = f"SOURCE-{engagement_id}-{plan_index:03d}-{week_index:03d}"
            conn.execute(
                """
                INSERT INTO time_entries (
                  snapshot_id, engagement_id, transaction_id, worker_name, worker_id, title,
                  worker_bu_du_cc, competency_center, entry_date, week_end_date, financial_period,
                  project_id, project_name, xref, phase_desc, task_desc, work_location,
                  billing_status, hours, fees_std_rate, fees_contract_rate, memo,
                  matched_phase_id, matched_team_member_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    engagement_id,
                    transaction_id,
                    member["name"],
                    f"SOURCE-{engagement_id:02d}-{member_id:03d}",
                    member["role"],
                    "",
                    "Source workbook",
                    week_end.isoformat(),
                    week_end.isoformat(),
                    week_end.strftime("%Y-%m"),
                    data["engagement_code"],
                    data["client_name"],
                    data["engagement_code"],
                    plan["phase_name"],
                    plan["phase_name"],
                    "",
                    "Source workbook",
                    hours,
                    round(hours * standard_rate, 2),
                    round(hours * contract_rate, 2),
                    "Exact weekly completed hours from the source workbook",
                    phase_id,
                    member_id,
                ),
            )
            rows_inserted += 1
        conn.execute("UPDATE weekly_snapshots SET row_count=? WHERE id=?", (rows_inserted, snapshot_id))


def seed() -> None:
    db_file = ROOT / "demo_seed.db"
    if db_file.exists():
        db_file.unlink()
    init_db(db_file)
    examples = sorted(
        path for path in WORKBOOK_DIR.glob("*.xls*")
        if not path.name.startswith("~$")
    )
    if not examples:
        raise FileNotFoundError(f"No example workbooks found under {WORKBOOK_DIR}")
    with connect(db_file) as conn:
        for path in examples:
            data = parse_example(path)
            engagement_id = insert_engagement(conn, data)
            phase_ids = [insert_phase(conn, engagement_id, phase, order) for order, phase in enumerate(data["phases"])]
            member_ids = [insert_member(conn, engagement_id, member) for member in data["team"]]
            insert_source_plans(conn, data, phase_ids, member_ids)
            insert_source_actuals(conn, engagement_id, data, phase_ids, member_ids)
            conn.execute(
                """
                INSERT INTO engagement_events (engagement_id, event_type, description, created_at)
                VALUES (?, 'seed_loaded', ?, ?)
                """,
                (
                    engagement_id,
                    f"Seeded from {path.name}",
                    now_iso(),
                ),
            )
    print(db_file)


if __name__ == "__main__":
    seed()
