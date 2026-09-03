from __future__ import annotations

import csv
import io
import re
import sqlite3
from datetime import datetime, timedelta
from typing import Any

from openpyxl import load_workbook

from calculations import as_float, money, variance_flag, week_monday
from db import get_app_settings


EXPECTED_COLUMNS = [
    "Transaction ID",
    "Worker ID",
    "Worker",
    "Title",
    "Worker BU DU CC",
    "Competency Center",
    "Date",
    "Week End Date",
    "Financial Period",
    "Project ID",
    "Project",
    "Xref",
    "Phase Desc",
    "Task Desc",
    "Work Loc",
    "Billing Status",
    "Hours",
    "Fees @ Std Rate",
    "Fees @ Contract Rate",
    "Memo",
]

DATE_COLUMNS = {"Date", "Week End Date"}
HEADER_MARKERS = {"Transaction ID", "Worker", "Hours", "Fees @ Contract Rate"}


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_match(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def excel_serial_to_iso(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    try:
        serial = float(text)
    except ValueError:
        return text
    return (datetime(1899, 12, 30) + timedelta(days=serial)).date().isoformat()


def parse_text_export(text: str) -> list[dict[str, Any]]:
    text = text.strip("\ufeff\r\n ")
    if not text:
        return []
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    raw_rows = [list(row) for row in reader]
    header_index, headers = _find_header_row(raw_rows)
    if header_index is None:
        return []
    return _records_from_rows(headers, raw_rows[header_index + 1 :])


def parse_xlsx_export(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    header_index, headers = _find_header_row(raw_rows)
    if header_index is None:
        return []
    return _records_from_rows(headers, raw_rows[header_index + 1 :])


def covered_period_from_text(text: str) -> tuple[str, str] | None:
    match = re.search(r"\bFrom\s*:\s*([0-9]{1,4}[/-][0-9]{1,2}[/-][0-9]{1,4})\s+to\s+"
                      r"([0-9]{1,4}[/-][0-9]{1,2}[/-][0-9]{1,4})", text, re.IGNORECASE)
    if not match:
        return None
    def parse_period_date(value: str) -> str:
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                pass
        return ""
    start = parse_period_date(match.group(1))
    end = parse_period_date(match.group(2))
    return (start, end) if start and end else None


def covered_period_from_xlsx(file_bytes: bytes) -> tuple[str, str] | None:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    preamble = " ".join(str(cell or "") for row in sheet.iter_rows(max_row=30, values_only=True) for cell in row)
    workbook.close()
    return covered_period_from_text(preamble)


def _find_header_row(rows: list[list[Any]]) -> tuple[int | None, list[str]]:
    best_index: int | None = None
    best_headers: list[str] = []
    best_score = 0
    for index, row in enumerate(rows):
        headers = [normalize_header(cell) for cell in row]
        present = {header for header in headers if header}
        score = sum(1 for column in EXPECTED_COLUMNS if column in present)
        if HEADER_MARKERS.issubset(present) and score > best_score:
            best_index = index
            best_headers = headers
            best_score = score
    return best_index, best_headers


def _records_from_rows(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows:
        if not any(value not in (None, "") for value in values):
            continue
        record = _row_from_values(headers, values)
        if _is_footer_or_non_entry(record):
            continue
        records.append(record)
    return records


def _is_footer_or_non_entry(record: dict[str, Any]) -> bool:
    transaction_id = str(record.get("Transaction ID", "")).strip()
    if not transaction_id:
        return True
    lowered = transaction_id.lower()
    if "summary" in lowered or lowered.startswith("overall"):
        return True
    return False


def _row_from_values(headers: list[str], values: list[Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        value = values[idx] if idx < len(values) else ""
        if header in DATE_COLUMNS:
            value = excel_serial_to_iso(value)
        record[header] = "" if value is None else value
    return record


def validate_columns(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return EXPECTED_COLUMNS
    present = set(rows[0].keys())
    return [column for column in EXPECTED_COLUMNS if column not in present]


def suggest_from_memo(memo: str, phases: list[dict[str, Any]]) -> dict[str, Any] | None:
    memo_norm = normalize_match(memo)
    if not memo_norm:
        return None
    for phase in phases:
        code = normalize_match(phase.get("phase_code"))
        name = normalize_match(phase.get("phase_name"))
        if (code and code in memo_norm) or (name and name in memo_norm):
            return {"phase_id": int(phase["id"]), "phase_name": phase.get("phase_name"),
                    "matched_text": phase.get("phase_code") if code and code in memo_norm else phase.get("phase_name")}
    return None


def preview_rows(
    conn: sqlite3.Connection, engagement_id: int, parsed_rows: list[dict[str, Any]],
    covered_period: tuple[str, str] | None = None,
) -> dict[str, Any]:
    engagement = conn.execute(
        "SELECT engagement_code, complexity_mode FROM engagements WHERE id = ?", (engagement_id,)
    ).fetchone()
    if engagement is None:
        raise ValueError("Engagement not found")

    team_rows = conn.execute(
        "SELECT id,name,is_active FROM team_members WHERE engagement_id=?", (engagement_id,)
    ).fetchall()
    # NOTE (see PRD §3.5): if a person is promoted mid-engagement (old team_members row for
    # the prior role, new row for the new role), Cognos exports one literal worker name
    # regardless of the title change, and this dict comprehension would silently last-win
    # between two rows sharing that name if both existed. Today this cannot actually happen:
    # team_members has UNIQUE(engagement_id, name COLLATE NOCASE) (schema.sql), which prevents
    # two rows sharing an exact name from ever coexisting. This ambiguity is real but
    # currently latent — it would only surface if that UNIQUE constraint were ever relaxed
    # (e.g. to support a promotion workaround). No fix in this pass; open product decision
    # (effective-dated rate vs. manual reassignment).
    team_by_name = {normalize_match(row["name"]): row for row in team_rows}
    covered_dates = sorted(
        value for value in (excel_serial_to_iso(row.get("Week End Date")) for row in parsed_rows) if value
    )
    covered_start = covered_period[0] if covered_period else (covered_dates[0] if covered_dates else None)
    covered_end = covered_period[1] if covered_period else (covered_dates[-1] if covered_dates else None)
    existing_rows = conn.execute(
        """SELECT * FROM time_entries WHERE engagement_id=?
        AND (? IS NULL OR week_end_date>=?) AND (? IS NULL OR week_end_date<=?)""",
        (engagement_id, covered_start, covered_start, covered_end, covered_end),
    ).fetchall()
    all_existing_rows = conn.execute(
        "SELECT * FROM time_entries WHERE engagement_id=? AND transaction_id IS NOT NULL", (engagement_id,)
    ).fetchall()
    existing_by_id = {str(row["transaction_id"]): row for row in all_existing_rows if row["transaction_id"]}
    phase_rows = conn.execute(
        "SELECT id, phase_name, phase_code FROM phases WHERE engagement_id=?", (engagement_id,)
    ).fetchall()
    exact_phases = {str(row["phase_code"] or "").strip(): int(row["id"]) for row in phase_rows
                    if str(row["phase_code"] or "").strip()}
    normalized_phases = {normalize_match(row["phase_code"]): int(row["id"]) for row in phase_rows
                         if normalize_match(row["phase_code"])}
    phase_names = {int(row["id"]): row["phase_name"] for row in phase_rows}
    sticky_rules = {int(row["team_member_id"]): int(row["phase_id"]) for row in conn.execute(
        "SELECT team_member_id, phase_id FROM allocation_rules WHERE engagement_id=?", (engagement_id,)
    ).fetchall()}
    staffing_weeks_by_member: dict[int, list[tuple[str, int]]] = {}
    for row in conn.execute(
        """SELECT ppw.team_member_id, ppw.phase_id, ppw.week_start_date FROM phase_person_weeks ppw
        JOIN phases p ON p.id = ppw.phase_id
        WHERE p.engagement_id=? AND (ppw.budgeted_hours>0 OR ppw.forecasted_hours>0)""", (engagement_id,)
    ).fetchall():
        staffing_weeks_by_member.setdefault(int(row["team_member_id"]), []).append(
            (row["week_start_date"], int(row["phase_id"]))
        )
    settings = get_app_settings(conn)
    weekly_existing = {
        (normalize_match(row["worker_name"]), week_monday(row["week_end_date"])): as_float(row["hours"])
        for row in conn.execute("""SELECT worker_name, week_end_date, SUM(hours) hours
        FROM time_entries WHERE engagement_id=? AND COALESCE(is_excluded,0)=0
        GROUP BY worker_name, week_end_date""", (engagement_id,))
    }

    batch_weekly: dict[tuple[str, str | None], float] = {}
    for record in parsed_rows:
        key = (normalize_match(record.get("Worker", "")), week_monday(excel_serial_to_iso(record.get("Week End Date"))))
        batch_weekly[key] = batch_weekly.get(key, 0) + as_float(record.get("Hours"))

    seen_ids: set[str] = set()
    rows: list[dict[str, Any]] = []
    summary = {"total": 0, "to_import": 0, "duplicates": 0, "flagged": 0,
               "rows_to_insert": 0, "rows_to_update": 0, "rows_to_remove": 0}
    for record in parsed_rows:
        transaction_id = str(record.get("Transaction ID", "")).strip()
        worker_name = str(record.get("Worker", "")).strip()
        normalized_worker_name = normalize_match(worker_name)
        member = team_by_name.get(normalized_worker_name)
        project_id = str(record.get("Project ID", "")).strip()
        hours = as_float(record.get("Hours"))
        phase_desc = str(record.get("Phase Desc", "")).strip()
        task_desc = str(record.get("Task Desc", "")).strip()
        monday = week_monday(excel_serial_to_iso(record.get("Week End Date")))
        allocation_method = None
        matched_phase_id = exact_phases.get(phase_desc) or normalized_phases.get(normalize_match(phase_desc))
        if matched_phase_id is not None:
            allocation_method = "direct_match"
        else:
            matched_phase_id = exact_phases.get(task_desc) or normalized_phases.get(normalize_match(task_desc))
            if matched_phase_id is not None:
                allocation_method = "task_match"
        if engagement["complexity_mode"] == "simple" and phase_rows:
            matched_phase_id = int(phase_rows[0]["id"])
            allocation_method = "direct_match"
        elif matched_phase_id is None and member is not None and bool(member["is_active"]):
            member_id = int(member["id"])
            rule_phase_id = sticky_rules.get(member_id)
            member_budgeted_phases = {phase_id for _, phase_id in staffing_weeks_by_member.get(member_id, [])}
            if rule_phase_id is not None:
                matched_phase_id = rule_phase_id
                allocation_method = "sticky_rule"
            elif len(member_budgeted_phases) == 1:
                # Budgeted for exactly one phase across the whole engagement — nothing to
                # decide, regardless of whether that budget falls near this entry's week.
                matched_phase_id = member_budgeted_phases.pop()
                allocation_method = "single_phase_budget"
            elif monday:
                # Only ambiguous cases (staffed on multiple concurrent phases) or workers with
                # no staffing signal at all should ever reach the exceptions queue — someone
                # staffed on exactly one phase that week has nothing to actually decide.
                window_start = (datetime.fromisoformat(monday).date() - timedelta(days=7)).isoformat()
                window_end = (datetime.fromisoformat(monday).date() + timedelta(days=7)).isoformat()
                candidates = {phase_id for week_start, phase_id in staffing_weeks_by_member.get(member_id, [])
                             if week_start and window_start <= week_start <= window_end}
                if len(candidates) == 1:
                    matched_phase_id = candidates.pop()
                    allocation_method = "staffing_match"
        flags: list[str] = []
        flag = None
        selectable = True
        included = True
        if not transaction_id or transaction_id in seen_ids:
            flag = "duplicate"
            selectable = False
            included = False
        elif hours == 0:
            flag = "zero_hours"
        elif worker_name and member is None:
            flag = "worker_unknown"
        elif member is not None and not bool(member["is_active"]):
            flag = "worker_unauthorized"
        elif project_id and project_id != engagement["engagement_code"]:
            flag = "project_mismatch"
        if flag:
            flags.append(flag)
        if matched_phase_id is None:
            flags.append("unmatched_phase")
            if flag is None:
                flag = "unmatched_phase"
        prior = None
        if monday:
            prior_day = (datetime.fromisoformat(monday).date() - timedelta(days=7)).isoformat()
            prior = weekly_existing.get((normalized_worker_name, prior_day))
        current_week = batch_weekly.get((normalized_worker_name, monday), hours)
        if variance_flag(current_week, prior, settings):
            flags.append("variance_flagged")
            if flag is None:
                flag = "variance_flagged"

        seen_ids.add(transaction_id)
        if flag == "duplicate":
            summary["duplicates"] += 1
        elif flag is not None:
            summary["flagged"] += 1
        if included:
            summary["to_import"] += 1
        summary["total"] += 1

        item = {
                "transaction_id": transaction_id,
                "worker_id": str(record.get("Worker ID", "")).strip(),
                "worker_name": worker_name,
                "normalized_worker_name": normalized_worker_name,
                "matched_team_member_id": int(member["id"]) if member is not None and bool(member["is_active"]) else None,
                "title": str(record.get("Title", "")).strip(),
                "worker_bu_du_cc": str(record.get("Worker BU DU CC", "")).strip(),
                "competency_center": str(record.get("Competency Center", "")).strip(),
                "entry_date": excel_serial_to_iso(record.get("Date")),
                "week_end_date": excel_serial_to_iso(record.get("Week End Date")),
                "financial_period": str(record.get("Financial Period", "")).strip(),
                "project_id": project_id,
                "project": str(record.get("Project", "")).strip(),
                "xref": str(record.get("Xref", "")).strip(),
                "phase_desc": phase_desc,
                "matched_phase_id": matched_phase_id,
                "matched_phase_name": phase_names.get(matched_phase_id),
                "allocation_method": allocation_method,
                "task_desc": str(record.get("Task Desc", "")).strip(),
                "work_location": str(record.get("Work Loc", "")).strip(),
                "billing_status": str(record.get("Billing Status", "")).strip(),
                "hours": hours,
                "fees_std_rate": money(record.get("Fees @ Std Rate")),
                "fees_contract_rate": money(record.get("Fees @ Contract Rate")),
                "memo": str(record.get("Memo", "")).strip(),
                "flag": flag,
                "flags": flags,
                "variance_flagged": "variance_flagged" in flags,
                "included": included,
                "selectable": selectable,
            }
        existing = existing_by_id.get(transaction_id)
        if existing is None:
            item["reconciliation_action"] = "insert"
            summary["rows_to_insert"] += 1
        else:
            compare_fields = ("worker_name", "entry_date", "week_end_date", "phase_desc", "hours",
                              "fees_std_rate", "fees_contract_rate", "matched_phase_id")
            before = {field: existing[field] for field in compare_fields}
            after = {field: item.get(field) for field in compare_fields}
            changed = any(str(before[field] if before[field] is not None else "") !=
                          str(after[field] if after[field] is not None else "") for field in compare_fields)
            item["reconciliation_action"] = "update" if changed else "unchanged"
            if changed:
                item["before"] = before
                item["after"] = after
                summary["rows_to_update"] += 1
        rows.append(item)

    incoming_ids = {row["transaction_id"] for row in rows if row["transaction_id"]}
    removals = [
        {"id": int(row["id"]), "transaction_id": row["transaction_id"],
         "worker_name": row["worker_name"], "week_end_date": row["week_end_date"],
         "hours": as_float(row["hours"]), "fees_contract_rate": money(row["fees_contract_rate"])}
        for row in existing_rows if row["transaction_id"] not in incoming_ids
    ]
    summary["rows_to_remove"] = len(removals)
    return {"rows": rows, "summary": summary, "covered_start_date": covered_start,
            "covered_end_date": covered_end, "rows_to_insert": summary["rows_to_insert"],
            "rows_to_update": summary["rows_to_update"], "rows_to_remove": removals}
