from __future__ import annotations

import html
import io
import re
import sqlite3
from copy import copy
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from calculations import engagement_metrics, phase_summary, team_summary
from db import row_to_dict, rows_to_dicts

NAVY = "011E41"
AMBER = "F5A800"
LIGHT = "F4F5F7"
WHITE = "FFFFFF"
TEXT = "333333"
BORDER = Side(style="thin", color="D9DDE3")


def filename_safe(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "Engagement"


def export_filename(engagement: dict[str, Any]) -> str:
    return f"{filename_safe(engagement['client_name'])}_{filename_safe(engagement['engagement_code'])}_{date.today().isoformat()}.xlsx"


def format_currency(value: Any) -> str:
    return f"${float(value or 0):,.2f}"


def format_hours(value: Any) -> str:
    return f"{float(value or 0):,.1f}"


def format_percent(value: Any) -> str:
    return "—" if value is None else f"{float(value):.1%}"


def engagement_payload(conn: sqlite3.Connection, engagement_id: int) -> dict[str, Any]:
    engagement = row_to_dict(conn.execute("SELECT * FROM engagements WHERE id=?", (engagement_id,)).fetchone())
    if not engagement:
        raise ValueError("Engagement not found")
    return {
        "engagement": engagement,
        "metrics": engagement_metrics(conn, engagement_id),
        "team": team_summary(conn, engagement_id),
        "phases": phase_summary(conn, engagement_id),
        "adjustments": rows_to_dicts(conn.execute("""SELECT a.*,p.phase_name FROM budget_adjustments a
            LEFT JOIN phases p ON p.id=a.phase_id WHERE a.engagement_id=? ORDER BY a.effective_date,a.id""", (engagement_id,)).fetchall()),
        "entries": rows_to_dicts(conn.execute("SELECT * FROM time_entries WHERE engagement_id=? ORDER BY week_end_date,worker_name,id", (engagement_id,)).fetchall()),
        "expenses": rows_to_dicts(conn.execute("SELECT * FROM expenses WHERE engagement_id=? ORDER BY incurred_date,id", (engagement_id,)).fetchall()),
        "revisions": rows_to_dicts(conn.execute("SELECT * FROM budget_revisions WHERE engagement_id=? ORDER BY revised_at,id", (engagement_id,)).fetchall()),
    }


def build_excel(conn: sqlite3.Connection, engagement_id: int) -> tuple[str, bytes]:
    payload = engagement_payload(conn, engagement_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Engagement Summary"
    _build_summary_sheet(ws, payload)
    _build_detail_sheet(wb, payload)
    _build_log_sheet(wb, "Adjustment Log", payload["adjustments"],
                     ["effective_date", "adjustment_type", "phase_name", "amount", "description"])
    _build_log_sheet(wb, "Expenses", payload["expenses"],
                     ["incurred_date", "expense_type", "phase_id", "amount", "description"])
    _build_log_sheet(wb, "Budget Revisions", payload["revisions"],
                     ["revised_at", "field_name", "old_value", "new_value", "reason"])
    stream = io.BytesIO()
    wb.save(stream)
    return export_filename(payload["engagement"]), stream.getvalue()


def _build_summary_sheet(ws, payload):
    e, m, team, phases = payload["engagement"], payload["metrics"], payload["team"], payload["phases"]
    ws.sheet_view.showGridLines = False
    ws.merge_cells("F2:I2")
    ws["F2"] = "Engagement Summary"
    ws["F2"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
    ws["F2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["F2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("K2:M2")
    ws["K2"] = "Budget position"
    ws["K2"].font = Font(name="Arial", bold=True, color=WHITE)
    ws["K2"].fill = PatternFill("solid", fgColor=NAVY)
    labels = [("B3", "Engagement Name"), ("B4", "Engagement Number"), ("B5", "Engagement Lead")]
    values = [("C3", e["client_name"]), ("C4", e["engagement_code"]), ("C5", e.get("engagement_lead") or "")]
    for cell, value in labels:
        ws[cell] = value
        ws[cell].font = Font(name="Arial", bold=True, color=TEXT)
    for cell, value in values:
        ws[cell] = value
    headers = ["Hours", "Internal/ Standard Fees", "Engagement Fees", "Realization Rate"]
    for index, value in enumerate(headers, 6):
        ws.cell(3, index, value)
    rows = [
        ("Total Budget", m["total_budgeted_hours"], sum(p["budgeted_std_fees"] for p in phases), m["total_budgeted_fees"],
         (m["signed_sow"]/sum(p["budgeted_std_fees"] for p in phases)) if sum(p["budgeted_std_fees"] for p in phases) else None),
        ("Actuals/Current Plan", sum(p["current_plan_hours"] for p in phases), m["fees_to_date_std"],
         sum(p["current_plan_eng_fees"] for p in phases), m["realization"]),
    ]
    for row_num, values_row in enumerate(rows, 4):
        for col, value in enumerate(values_row, 5):
            ws.cell(row_num, col, value)
    ws["E6"] = "Variance From Plan"
    ws["F6"] = rows[1][1]-rows[0][1]
    ws["G6"] = (rows[1][2]-rows[0][2])/rows[0][2] if rows[0][2] else None
    ws["H6"] = (rows[1][3]-rows[0][3])/rows[0][3] if rows[0][3] else None
    ws.merge_cells("F7:G7")
    ws["F7"] = "Potential Change Order Amount →"
    ws["H7"] = max(0, rows[1][3]-rows[0][3])
    ws["K3"], ws["L3"], ws["M3"] = "SOW Fees", "Current/Actual Engagement Fees", "Variance"
    ws["K4"], ws["L4"], ws["M4"] = m["signed_sow"], rows[1][3], m["signed_sow"]-rows[1][3]
    _summary_tables(ws, team, phases)
    _style_summary(ws)


def _summary_tables(ws, team, phases):
    ws.merge_cells("B10:G10")
    ws["B10"] = "Engagement Team Summary"
    ws.merge_cells("I10:O10")
    ws["I10"] = "Effort Summary"
    team_headers = ["Name", "Project Role", "Hours (Budget)", "Hours (Actual/Current)",
                    "Engagement Fees (Budget)", "Engagement Fees (Current Plan)"]
    phase_headers = ["Phase or Segment", "Budgeted Hours", "Actual/Current Plan Hours",
                     "Hours Variance", "Engagement Budget", "Engagement Fees Planned", "Over/Under Budget"]
    for col, value in enumerate(team_headers, 2):
        ws.cell(11, col, value)
    for col, value in enumerate(phase_headers, 9):
        ws.cell(11, col, value)
    for row_num, member in enumerate(team, 12):
        values = [member["name"], member.get("role") or "", member["budgeted_hours"],
                  member["hours_to_date"], member["budgeted_eng_fees"], member["actual_eng_fees"]]
        for col, value in enumerate(values, 2):
            ws.cell(row_num, col, value)
    for row_num, phase in enumerate(phases, 12):
        values = [phase["phase_name"], phase["budgeted_hours"], phase["current_plan_hours"],
                  phase["current_plan_hours"]-phase["budgeted_hours"], phase["effective_sow"],
                  phase["current_plan_eng_fees"], phase["current_plan_eng_fees"]-phase["effective_sow"]]
        for col, value in enumerate(values, 9):
            ws.cell(row_num, col, value)
    total_row = 12 + max(len(team), len(phases))
    ws.cell(total_row, 2, "Total")
    ws.cell(total_row, 9, "Total")
    for col in range(4, 8):
        ws.cell(total_row, col, f"=SUM({get_column_letter(col)}12:{get_column_letter(col)}{total_row-1})")
    for col in range(10, 16):
        ws.cell(total_row, col, f"=SUM({get_column_letter(col)}12:{get_column_letter(col)}{total_row-1})")
    for cell in ws[total_row]:
        cell.font = Font(name="Arial", bold=True)


def _style_summary(ws):
    widths = {"B": 24, "C": 23, "D": 15, "E": 20, "F": 22, "G": 22,
              "H": 18, "I": 30, "J": 16, "K": 22, "L": 18, "M": 20, "N": 24, "O": 20}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=15):
        for cell in row:
            font = copy(cell.font)
            font.name = "Arial"
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_num in (3, 11):
        for cell in ws[row_num][1:15]:
            if cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
                cell.font = Font(name="Arial", bold=True, color=TEXT)
                cell.border = Border(bottom=BORDER)
    for cell in (ws["B10"], ws["I10"]):
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
    for row in ws.iter_rows(min_row=12, max_row=ws.max_row, min_col=4, max_col=15):
        for cell in row:
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00' if cell.column in {6,7,13,14,15} else '#,##0.0'
    for cell in ("G4", "H4", "G5", "H5", "H7", "K4", "L4", "M4"):
        ws[cell].number_format = '$#,##0.00;[Red]-$#,##0.00'
    for cell in ("I4", "I5", "G6", "H6"):
        ws[cell].number_format = '0.0%'
    ws.freeze_panes = "B11"
    ws.print_area = f"B2:O{ws.max_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _build_detail_sheet(wb, payload):
    ws = wb.create_sheet("Weekly Detail")
    headers = ["transaction_id", "worker_id", "worker_name", "title", "worker_bu_du_cc",
               "competency_center", "entry_date", "week_end_date", "financial_period",
               "project_id", "project_name", "xref", "phase_desc", "task_desc",
               "work_location", "billing_status", "hours", "fees_std_rate",
               "fees_contract_rate", "memo", "matched_phase_id"]
    ws.append([header.replace("_", " ").title() for header in headers])
    for entry in payload["entries"]:
        ws.append([entry.get(header) for header in headers])
    _style_tabular(ws)
    for row in range(2, ws.max_row+1):
        ws.cell(row, 17).number_format = '#,##0.00'
        ws.cell(row, 18).number_format = '$#,##0.00'
        ws.cell(row, 19).number_format = '$#,##0.00'


def _build_log_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append([field.replace("_", " ").title() for field in fields])
    for item in rows:
        ws.append([item.get(field) for field in fields])
    _style_tabular(ws)


def _style_tabular(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column in range(1, ws.max_column+1):
        letter = get_column_letter(column)
        width = max((len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, 200)+1)), default=10)
        ws.column_dimensions[letter].width = min(max(width+2, 12), 32)


def build_html_report(conn: sqlite3.Connection, engagement_id: int, narrative: str = "") -> str:
    p = engagement_payload(conn, engagement_id)
    e, m = p["engagement"], p["metrics"]
    team_rows = "".join(f"<tr><td>{html.escape(str(x['name']))}</td><td>{html.escape(str(x.get('role') or ''))}</td>"
        f"<td>{format_hours(x['budgeted_hours'])}</td><td>{format_hours(x['hours_to_date'])}</td>"
        f"<td>{format_currency(x['budgeted_eng_fees'])}</td><td>{format_currency(x['actual_eng_fees'])}</td></tr>" for x in p["team"])
    phase_rows = "".join(f"<tr><td>{html.escape(str(x['phase_name']))}</td><td>{format_hours(x['budgeted_hours'])}</td>"
        f"<td>{format_hours(x['current_plan_hours'])}</td><td>{format_hours(x['current_plan_hours']-x['budgeted_hours'])}</td>"
        f"<td>{format_currency(x['effective_sow'])}</td><td>{format_currency(x['current_plan_eng_fees'])}</td>"
        f"<td>{format_currency(x['current_plan_eng_fees']-x['effective_sow'])}</td></tr>" for x in p["phases"])
    return f"""<!doctype html><html><head><meta charset='utf-8'><title>Engagement Summary</title>
    <style>{_report_css()}</style></head><body><main>
    <header><div><div class='eyebrow'>Engagement Summary</div><h1>{html.escape(e['client_name'])}</h1>
    <p>{html.escape(e['engagement_code'])} · {html.escape(str(e.get('engagement_lead') or ''))}</p></div>
    <div class='status'>{html.escape(m['status'])}</div></header>
    <section class='summary'><table><thead><tr><th></th><th>Hours</th><th>Internal/Standard Fees</th><th>Engagement Fees</th><th>Realization</th></tr></thead>
    <tbody><tr><th>Total Budget</th><td>{format_hours(m['total_budgeted_hours'])}</td><td>{format_currency(sum(x['budgeted_std_fees'] for x in p['phases']))}</td><td>{format_currency(m['total_budgeted_fees'])}</td><td>—</td></tr>
    <tr><th>Actuals/Current Plan</th><td>{format_hours(sum(x['current_plan_hours'] for x in p['phases']))}</td><td>{format_currency(m['fees_to_date_std'])}</td><td>{format_currency(sum(x['current_plan_eng_fees'] for x in p['phases']))}</td><td>{format_percent(m['realization'])}</td></tr></tbody></table>
    <aside><span>Signed SOW</span><strong>{format_currency(m['signed_sow'])}</strong><span>Effective budget</span><strong>{format_currency(m['effective_sow'])}</strong></aside></section>
    <section class='tables'><div><h2>Engagement Team Summary</h2><table><thead><tr><th>Name</th><th>Project Role</th><th>Hours Budget</th><th>Hours Actual</th><th>Fees Budget</th><th>Fees Current</th></tr></thead><tbody>{team_rows}</tbody></table></div>
    <div><h2>Effort Summary</h2><table><thead><tr><th>Phase</th><th>Budget Hours</th><th>Actual Hours</th><th>Variance</th><th>Budget</th><th>Fees Planned</th><th>Over/Under</th></tr></thead><tbody>{phase_rows}</tbody></table></div></section>
    {f"<section class='narrative'><h2>Status narrative</h2><p>{html.escape(narrative)}</p></section>" if narrative else ''}
    <footer>Generated {date.today().isoformat()} · Engagement Budget Tracker</footer>
    </main></body></html>"""


def _report_css():
    return """@page{size:landscape;margin:12mm}*{box-sizing:border-box}body{margin:0;background:#fff;color:#333;font:12px Arial,sans-serif}main{max-width:1400px;margin:auto}header{display:flex;justify-content:space-between;align-items:end;padding:24px 28px;background:#011E41;color:#fff;border-top:7px solid #F5A800}.eyebrow{text-transform:uppercase;letter-spacing:.14em;color:#F5A800;font-weight:700}h1{margin:6px 0 2px;font-size:28px}header p{margin:0;color:#d7e1ee}.status{border:1px solid #F5A800;padding:8px 12px;font-weight:700}.summary{display:grid;grid-template-columns:1fr 240px;gap:20px;margin:24px 0}.summary aside{background:#f4f5f7;padding:18px;display:grid;gap:7px}.summary aside strong{font-size:18px;color:#011E41;margin-bottom:8px}.tables{display:grid;grid-template-columns:1fr 1.15fr;gap:18px}h2{font-size:15px;color:#011E41;border-bottom:3px solid #F5A800;padding-bottom:6px}table{width:100%;border-collapse:collapse}th,td{padding:7px 8px;border-bottom:1px solid #d9dde3;text-align:right}th:first-child,td:first-child{text-align:left}thead th{background:#011E41;color:#fff;font-size:10px}.narrative{break-before:page;margin-top:24px}footer{margin-top:28px;border-top:1px solid #d9dde3;padding-top:8px;color:#666}@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact}}"""
